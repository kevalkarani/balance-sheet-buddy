"""
Output generation functions for Balance Sheet Buddy.
Creates formatted Excel files and reports from Claude's analysis.
"""

import pandas as pd
import io
import re
from typing import Dict, List, Tuple
from datetime import datetime


def parse_claude_table_response(response_text: str) -> pd.DataFrame:
    """
    Parse Claude's table response into a DataFrame.
    Expects format: Account | Balance_Type | Amount | Category | Subcategory | Commentary | Status
    """
    lines = response_text.strip().split('\n')

    # Find table start (look for header with |)
    table_start = None
    for i, line in enumerate(lines):
        if '|' in line and any(keyword in line.lower() for keyword in ['account', 'balance', 'status', 'commentary']):
            table_start = i
            break

    if table_start is None:
        # No table found, return empty DataFrame
        return pd.DataFrame()

    # Extract table lines (skip separator lines with only -, |, and spaces)
    table_lines = []
    for line in lines[table_start:]:
        # Skip separator lines but keep data lines
        if '|' in line and not re.match(r'^[\s\-|:]+$', line):
            # Clean up the line - remove leading/trailing pipes
            clean_line = line.strip()
            if clean_line.startswith('|'):
                clean_line = clean_line[1:]
            if clean_line.endswith('|'):
                clean_line = clean_line[:-1]
            table_lines.append(clean_line)

    if len(table_lines) < 2:  # Need at least header and one row
        return pd.DataFrame()

    # Parse headers
    headers = [h.strip() for h in table_lines[0].split('|')]
    headers = [h for h in headers if h]  # Remove empty strings

    # Parse data rows
    rows = []
    for line in table_lines[1:]:
        cells = [c.strip() for c in line.split('|')]
        # Remove empty cells
        cells = [c if c else '' for c in cells]

        if len(cells) >= len(headers):
            rows.append(cells[:len(headers)])
        elif len(cells) > 0:
            # Pad with empty strings if row is shorter
            while len(cells) < len(headers):
                cells.append('')
            rows.append(cells)

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows, columns=headers)

    # Standardize column names to match expected format
    column_mapping = {}
    for col in df.columns:
        col_lower = col.lower()
        if 'account' in col_lower:
            column_mapping[col] = 'Account'
        elif 'balance' in col_lower and 'type' in col_lower:
            column_mapping[col] = 'Balance_Type'
        elif 'amount' in col_lower:
            column_mapping[col] = 'Amount'
        elif 'category' in col_lower and 'sub' not in col_lower:
            column_mapping[col] = 'Category'
        elif 'subcategory' in col_lower or 'sub-category' in col_lower or 'sub category' in col_lower:
            column_mapping[col] = 'Subcategory'
        elif 'comment' in col_lower:
            column_mapping[col] = 'Commentary'
        elif 'status' in col_lower:
            column_mapping[col] = 'Status'

    if column_mapping:
        df = df.rename(columns=column_mapping)

    return df


def create_classification_excel(analysis_text: str, tb_df: pd.DataFrame) -> io.BytesIO:
    """
    Create Output A - Classification View Excel file.

    Columns:
    - Account No - Name
    - Debit/Credit indicator
    - Amount
    - Category
    - Subcategory
    - Status (PASS or MISMATCH)
    - Classification commentary
    """
    # Try to parse Claude's response as a table
    result_df = parse_claude_table_response(analysis_text)

    # If parsing failed or incomplete, create from trial balance with inferred status
    if result_df.empty or len(result_df) < len(tb_df) or 'Status' not in result_df.columns:
        result_df = tb_df.copy()

        # Add missing columns
        if 'Balance_Type' not in result_df.columns:
            result_df['Balance_Type'] = result_df.apply(
                lambda row: 'Debit' if row['Debit'] > 0 else 'Credit' if row['Credit'] > 0 else 'Zero',
                axis=1
            )

        if 'Amount' not in result_df.columns:
            result_df['Amount'] = result_df.apply(
                lambda row: row['Debit'] if row['Debit'] > 0 else row['Credit'],
                axis=1
            )

        # Infer Status based on category and balance type
        if 'Status' not in result_df.columns:
            def infer_status(row):
                category = str(row.get('Category', '')).lower()
                balance_type = row.get('Balance_Type', '')

                # Apply framework rules
                if 'asset' in category and 'contra' not in category:
                    return 'PASS' if balance_type == 'Debit' else 'MISMATCH'
                elif 'liability' in category or 'equity' in category:
                    return 'PASS' if balance_type == 'Credit' else 'MISMATCH'
                elif 'clearing' in category:
                    return 'PASS' if balance_type == 'Zero' else 'MISMATCH'
                else:
                    return 'PASS'  # Default for unmapped

            result_df['Status'] = result_df.apply(infer_status, axis=1)

        # Generate commentary based on status
        if 'Commentary' not in result_df.columns or result_df['Commentary'].str.contains('pending', case=False, na=True).any():
            def generate_commentary(row):
                status = row.get('Status', '')
                category = row.get('Category', '')
                balance_type = row.get('Balance_Type', '')

                if status == 'PASS':
                    return f"{category} with {balance_type} balance - Correct"
                elif status == 'MISMATCH':
                    return f"{category} with {balance_type} balance - Incorrect (should be opposite)"
                else:
                    return f"{category} - Review required"

            result_df['Commentary'] = result_df.apply(generate_commentary, axis=1)

    # Ensure all required columns exist
    required_cols = ['Account', 'Balance_Type', 'Amount', 'Category', 'Subcategory', 'Status', 'Commentary']
    for col in required_cols:
        if col not in result_df.columns:
            result_df[col] = ''

    # Reorder columns - Status comes BEFORE Commentary
    output_df = result_df[required_cols]

    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        output_df.to_excel(writer, index=False, sheet_name='Classification View')

        # Get the worksheet
        worksheet = writer.sheets['Classification View']

        # Format headers (bold)
        for cell in worksheet[1]:
            cell.font = cell.font.copy(bold=True)

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Color code status column
        for row in range(2, len(output_df) + 2):
            status_cell = worksheet.cell(row=row, column=7)  # Status is column 7
            if status_cell.value == 'MISMATCH':
                from openpyxl.styles import PatternFill
                status_cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            elif status_cell.value == 'PASS':
                from openpyxl.styles import PatternFill
                status_cell.fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')

    output.seek(0)
    return output


def create_classification_excel_from_df(df: pd.DataFrame) -> io.BytesIO:
    """
    Create Excel output directly from the classification DataFrame.
    This ensures Excel matches exactly what's displayed on screen.
    """
    from openpyxl.styles import PatternFill, Font

    # Ensure required columns exist
    required_cols = ['Account', 'Balance_Type', 'Amount', 'Category', 'Subcategory', 'Status', 'Commentary']

    # Add missing columns if needed
    if 'Commentary' not in df.columns:
        def generate_commentary(row):
            status = row.get('Status', '')
            category = row.get('Category', '')
            balance_type = row.get('Balance_Type', '')

            if status == 'PASS':
                return f"{category} with {balance_type} balance - Correct"
            elif status == 'MISMATCH':
                return f"{category} with {balance_type} balance - Should be opposite"
            else:
                return f"{category} - Review required"

        df = df.copy()
        df['Commentary'] = df.apply(generate_commentary, axis=1)

    # Select and order columns
    output_cols = [col for col in required_cols if col in df.columns]
    output_df = df[output_cols].copy()

    # Create Excel file
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        output_df.to_excel(writer, index=False, sheet_name='Classification View')

        worksheet = writer.sheets['Classification View']

        # Format headers
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Color code Status column
        status_col_idx = output_cols.index('Status') + 1 if 'Status' in output_cols else None
        if status_col_idx:
            for row in range(2, len(output_df) + 2):
                status_cell = worksheet.cell(row=row, column=status_col_idx)
                if status_cell.value == 'MISMATCH':
                    status_cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                elif status_cell.value == 'PASS':
                    status_cell.fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')

    output.seek(0)
    return output


def format_reconciliation_report(analysis_text: str) -> str:
    """
    Format Output B - Account-Level Reconciliation report as HTML.
    """
    # Clean up the text
    analysis_text = analysis_text.strip()

    # Convert to HTML with basic formatting
    html = f"""
    <div style="font-family: Arial, sans-serif; padding: 20px;">
        <h2>Output B: Account-Level Reconciliation</h2>
        <div style="white-space: pre-wrap; line-height: 1.6;">
            {analysis_text}
        </div>
    </div>
    """
    return html


def format_executive_summary(analysis_text: str) -> str:
    """
    Format Output C - Executive Summary as HTML.
    """
    # Extract executive summary section if present
    summary_match = re.search(r'OUTPUT C.*?EXECUTIVE SUMMARY.*?$(.*)', analysis_text, re.IGNORECASE | re.DOTALL)
    if summary_match:
        summary_text = summary_match.group(1).strip()
    else:
        summary_text = analysis_text.strip()

    html = f"""
    <div style="font-family: Arial, sans-serif; padding: 20px;">
        <h2>Output C: Executive Summary</h2>
        <div style="white-space: pre-wrap; line-height: 1.6;">
            {summary_text}
        </div>
    </div>
    """
    return html


def create_combined_report(classification_text: str, reconciliation_text: str = None, summary_text: str = None) -> str:
    """
    Create a combined HTML report with all outputs.
    """
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    html_parts = [f"""
    <html>
    <head>
        <title>Balance Sheet Buddy - Analysis Report</title>
        <style>
            body {{
                font-family: Arial, sans-serif;
                margin: 40px;
                background-color: #f5f5f5;
            }}
            .header {{
                background-color: #1f77b4;
                color: white;
                padding: 20px;
                border-radius: 5px;
                margin-bottom: 20px;
            }}
            .section {{
                background-color: white;
                padding: 20px;
                margin-bottom: 20px;
                border-radius: 5px;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            }}
            h2 {{
                color: #1f77b4;
                border-bottom: 2px solid #1f77b4;
                padding-bottom: 10px;
            }}
            .timestamp {{
                color: #666;
                font-size: 0.9em;
            }}
            .status-pass {{
                color: green;
                font-weight: bold;
            }}
            .status-mismatch {{
                color: red;
                font-weight: bold;
            }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>Balance Sheet Buddy - Analysis Report</h1>
            <p class="timestamp">Generated: {timestamp}</p>
        </div>
    """]

    # Add classification output
    if classification_text:
        html_parts.append(f"""
        <div class="section">
            <h2>Classification Analysis</h2>
            <pre style="white-space: pre-wrap; line-height: 1.6;">{classification_text}</pre>
        </div>
        """)

    # Add reconciliation output
    if reconciliation_text:
        html_parts.append(f"""
        <div class="section">
            <h2>Account-Level Reconciliation</h2>
            <pre style="white-space: pre-wrap; line-height: 1.6;">{reconciliation_text}</pre>
        </div>
        """)

    # Add executive summary
    if summary_text:
        html_parts.append(f"""
        <div class="section">
            <h2>Executive Summary</h2>
            <pre style="white-space: pre-wrap; line-height: 1.6;">{summary_text}</pre>
        </div>
        """)

    html_parts.append("""
    </body>
    </html>
    """)

    return "".join(html_parts)


def extract_summary_stats(classification_text: str, df: pd.DataFrame = None) -> Dict:
    """
    Extract summary statistics from classification analysis.
    Returns dict with counts of PASS, MISMATCH, total accounts, category breakdowns, etc.
    """
    stats = {
        'total_accounts': 0,
        'pass_count': 0,
        'mismatch_count': 0,
        'unmapped_count': 0,
        'by_category': {},
        'by_status': {},
        'total_debit': 0,
        'total_credit': 0
    }

    # If DataFrame is provided, extract from it (more accurate)
    if df is not None and not df.empty:
        stats['total_accounts'] = len(df)

        if 'Status' in df.columns:
            # Filter out NaN values before counting
            status_series = df['Status'].dropna().astype(str).str.strip()
            stats['pass_count'] = len(status_series[status_series.str.upper() == 'PASS'])
            stats['mismatch_count'] = len(status_series[status_series.str.upper() == 'MISMATCH'])

            # Count accounts with other/empty status
            all_status = df['Status'].fillna('').astype(str).str.strip().str.upper()
            stats['other_status_count'] = len(df[~all_status.isin(['PASS', 'MISMATCH', ''])])

        if 'Category' in df.columns:
            stats['unmapped_count'] = len(df[df['Category'] == 'Unmapped'])
            # Category breakdown
            category_counts = df['Category'].value_counts().to_dict()
            stats['by_category'] = category_counts

        if 'Status' in df.columns and 'Category' in df.columns:
            # Status by category
            for category in df['Category'].unique():
                cat_df = df[df['Category'] == category]
                status_series = cat_df['Status'].dropna().astype(str)
                stats['by_status'][category] = {
                    'total': len(cat_df),
                    'pass': len(status_series[status_series.str.upper() == 'PASS']),
                    'mismatch': len(status_series[status_series.str.upper() == 'MISMATCH'])
                }

        # Calculate totals
        if 'Debit' in df.columns:
            stats['total_debit'] = df['Debit'].sum()
        if 'Credit' in df.columns:
            stats['total_credit'] = df['Credit'].sum()

    else:
        # Fallback to text parsing
        stats['pass_count'] = len(re.findall(r'\bPASS\b', classification_text, re.IGNORECASE))
        stats['mismatch_count'] = len(re.findall(r'\bMISMATCH\b', classification_text, re.IGNORECASE))
        stats['unmapped_count'] = len(re.findall(r'\bUnmapped\b', classification_text, re.IGNORECASE))
        stats['total_accounts'] = stats['pass_count'] + stats['mismatch_count']

    return stats


def create_summary_text(stats: Dict, df: pd.DataFrame = None) -> str:
    """
    Create a detailed text summary of analysis results.
    """
    total = stats.get('total_accounts', 0)
    passed = stats.get('pass_count', 0)
    mismatched = stats.get('mismatch_count', 0)
    unmapped = stats.get('unmapped_count', 0)

    pass_pct = (passed/total*100) if total > 0 else 0
    mismatch_pct = (mismatched/total*100) if total > 0 else 0

    other_status = stats.get('other_status_count', 0)

    lines = [
        "📊 ANALYSIS SUMMARY",
        "=" * 60,
        "",
        f"📋 Total Accounts Analyzed: {total}",
        f"✅ PASS: {passed} ({pass_pct:.1f}%)",
        f"❌ MISMATCH: {mismatched} ({mismatch_pct:.1f}%)",
        f"⚠️  Unmapped Accounts: {unmapped}",
    ]

    # Show other status count if exists
    if other_status > 0:
        lines.append(f"ℹ️  Other Status: {other_status}")

    # Validation check
    accounted_for = passed + mismatched + other_status
    if accounted_for != total:
        missing = total - accounted_for
        lines.append(f"⚠️  WARNING: {missing} accounts with empty/missing Status")

    lines.append("")

    # Add balance totals if available
    if 'total_debit' in stats and 'total_credit' in stats:
        total_debit = stats['total_debit']
        total_credit = stats['total_credit']
        balance_diff = abs(total_debit - total_credit)
        is_balanced = balance_diff < 0.01  # Allow for rounding errors

        lines.extend([
            "💰 TRIAL BALANCE TOTALS:",
            f"   Total Debits:  ${total_debit:,.2f}",
            f"   Total Credits: ${total_credit:,.2f}",
            f"   Difference:    ${balance_diff:,.2f}",
        ])

        if is_balanced:
            lines.append(f"   ✅ Trial Balance is BALANCED")
        else:
            lines.append(f"   ❌ Trial Balance OUT OF BALANCE by ${balance_diff:,.2f}")
        lines.append("")

    # List mismatched accounts if available
    if mismatched > 0 and df is not None and 'Status' in df.columns:
        mismatch_df = df[df['Status'].str.upper() == 'MISMATCH']
        if not mismatch_df.empty:
            lines.extend([
                "⚠️  ACCOUNTS REQUIRING REVIEW:",
                "-" * 60
            ])
            for idx, row in mismatch_df.iterrows():
                account = row.get('Account', 'Unknown')
                category = row.get('Category', 'N/A')
                amount = row.get('Amount', 0)
                # Convert amount to float if it's a string
                try:
                    amount_float = float(amount) if amount else 0.0
                except (ValueError, TypeError):
                    amount_float = 0.0
                lines.append(f"   • {account} ({category}) - ${amount_float:,.2f}")
            lines.append("")

    # Add category breakdown if available
    if stats.get('by_category'):
        lines.extend([
            "📂 BREAKDOWN BY CATEGORY:",
            "-" * 60
        ])
        for category, count in sorted(stats['by_category'].items(), key=lambda x: x[1], reverse=True):
            pct = (count/total*100) if total > 0 else 0

            # Get status breakdown for this category
            status_info = ""
            if category in stats.get('by_status', {}):
                cat_stats = stats['by_status'][category]
                status_info = f" (✅ {cat_stats['pass']} | ❌ {cat_stats['mismatch']})"

            lines.append(f"   {category:20s}: {count:3d} accounts ({pct:5.1f}%){status_info}")
        lines.append("")

    # Overall status
    if mismatched == 0:
        lines.extend([
            "=" * 60,
            "✨ ALL ACCOUNTS VALIDATED SUCCESSFULLY!",
            "=" * 60
        ])
    else:
        lines.extend([
            "=" * 60,
            f"⚠️  ATTENTION REQUIRED: {mismatched} account(s) need review",
            "=" * 60
        ])

    return "\n".join(lines)
